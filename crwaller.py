#-*- coding: utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook, load_workbook
import os
import time



filename = "cosmicgirl.xlsx"
pageflag = 0

class DcInsideParser:
    def __init__(self, phantom_js, gall_id):
        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        options.add_argument('window-size=1920x1080')
        options.add_argument("disable-gpu")
        self._browser = webdriver.Chrome(executable_path=phantom_js, chrome_options=options)
        self.visitdict = {}
        self._gall_id = gall_id
        self._url = 'http://gall.dcinside.com'
        self._page_no = None
        self._posts = []
        self.pageflag = pageflag
        self.sdate = ''
        self.edate = ''
        self.pre = ''
        self.suf = ''

    def set_page_no(self, page_no):
        self._page_no = page_no

    def set_date(self, sdate, edate):
        self.sdate = sdate
        self.edate = edate

    def set_filter(self, pre, suf):
        self.pre = pre
        self.suf = suf

    def load_document(self, minor):
        if minor == False:
            url = self._url + '/board/lists/?id=' + self._gall_id
        else:
            url = self._url + '/mgallery/board/lists/?id=' + self._gall_id
        if self._page_no is not None:
            url += '&page=' + str(self._page_no)
        self._browser.get(url)

    def creatExcelfile(self, filename, dir):
        if os.path.exists(filename):
            wb = load_workbook(filename)
            now = time.localtime()
            s = "%04d-%02d-%02d (%02d-%02d-%02d)" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
            ws = wb.create_sheet(title=s)
            datelist = list(self.visitdict.keys())
            for i in range(len(datelist)):
                ws.cell(row=1, column=i+1).value = datelist[i]
                for j in range(len(self.visitdict[datelist[i]])):
                    ws.cell(row=j+2, column=i+1).value = self.visitdict[datelist[i]][j]
            wb.save(dir + "/"+ filename)
        else:
            wb = Workbook()
            now = time.localtime()
            s = "%04d-%02d-%02d (%02d-%02d-%02d)" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
            ws = wb.create_sheet(title=s)
            datelist = list(self.visitdict.keys())
            for i in range(len(datelist)):
                ws.cell(row=1, column=i+1).value = datelist[i]
                for j in range(len(self.visitdict[datelist[i]])):
                    ws.cell(row=j+2, column=i+1).value = self.visitdict[datelist[i]][j]

            wb.save(dir + "/"+ filename)

    def load_posts(self):
        self.pageflag = 0
        countexecption  = 0;

        lines = self._browser.find_elements_by_class_name("ub-content")
        for line in lines:
            text_no = line.find_element_by_class_name("gall_subject").text
            #print(text_no)
            writer = line.find_element_by_class_name("gall_writer").text
            #print(writer)
            post = line.find_element_by_class_name('gall_tit').text
            #print(post)
            date = line.find_element_by_class_name('gall_date').text
            #print(date)
            mdate = str(date).replace(".", "")
            #mdate = str(date).replace("/", "")
            if ":" in mdate:
                now = time.localtime()
                mdate = "%02d%02d" % (now.tm_mon, now.tm_mday)
            #print (mdate)
            if text_no in '공지' or text_no in '이슈' or text_no in '설문' or text_no in 'AD':
                continue
            else:
                mdate = "2019" + str(mdate)
                if int(mdate) >= int(self.edate): #확인
                    self.pageflag = 1
                #if '[' in post and '갤]' in post:   old version
                if self.pre in post and self.suf in post:
                    if int(mdate) <= int(self.sdate) and int(mdate) >= int(self.edate):
                        #print (mdate)
                        #findex = str(post).find(u'[')
                        #eindex = str(post).find(u']')
                        findex = str(post).find(self.pre)
                        eindex = len(self.pre) + str(post[len(self.pre):]).find(self.suf)
                        gallname = post[findex+len(self.pre):eindex]
                        gallnamewithid = gallname+"("+writer+")"
                        #print(gallnamewithid)
                        try:
                            val = int(gallname)
                        except ValueError:
                            if mdate in self.visitdict.keys():
                                self.visitdict[mdate].append(gallnamewithid)
                            else:
                                self.visitdict[mdate] = [gallnamewithid]

            '''   
            else:
                link = post.find_elements_by_tag_name('a')[0]
                self._posts.append(link.get_attribute('href'))
                print (text_no)
            '''

    def load_post(self):
        # 실제로 코멘트 데이터가 들어간 리스트
        comment_data_list = []
        for post in self._posts:
            print ('load post = ' + post)
            self._browser.get(post)
            try:
                wait = WebDriverWait(self._browser, 10)
                replies = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'reply_line')))
                for reply in replies:
                    comment_data = {'writer': reply.find_element_by_class_name('user_layer').text,
                                    'content': reply.find_element_by_class_name('reply').text,
                                    'date': reply.find_element_by_class_name('retime').text}
                    comment_data_list.append(comment_data)
                    print (comment_data['writer'] + " / " + comment_data['content'])
            except TimeoutException:
                continue

